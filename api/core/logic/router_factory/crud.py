import json
import csv
import io
import logging
from pathlib import Path
from fastapi import APIRouter, Depends, Query, Request, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, text, func
from typing import List, Type, Any, Optional
from pydantic import ValidationError

from .helpers import (
    _apply_scope_filters, _inject_scope_payload, _check_scope_ownership, 
    _save_children, _broadcast_update, _trigger_invalidation,
    _get_scope
)
from ...lib.database import get_db
from ...logic.permissions import check_permissions
from ...exceptions import ValidationException, ResourceNotFoundException
from ...response import ok

MAX_PER_PAGE = 200
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_PREVIEW_ROWS = 200
ALLOWED_IMPORT_TYPES = {
    "text/csv",
    "application/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
ALLOWED_IMPORT_EXTENSIONS = {".csv", ".xlsx"}


# gpt-5
def _normalize_import_mapping(mapping: Optional[dict]) -> dict[str, str]:
    if mapping is None:
        return {}
    if not isinstance(mapping, dict):
        raise ValidationException("Invalid mapping format")
    return {
        str(source): str(target)
        for source, target in mapping.items()
        if source and target
    }


# gpt-5
def _clean_import_row(row: dict[str, Any]) -> dict[str, Any]:
    cleaned_row: dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized_key = str(key).strip()
        if not normalized_key:
            continue
        if isinstance(value, str):
            value = value.strip()
        cleaned_row[normalized_key] = None if value == "" else value
    return cleaned_row


# gpt-5
def _apply_import_mapping(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    if not mapping:
        return row
    mapped_row: dict[str, Any] = {}
    for source_key, value in row.items():
        target_key = mapping.get(source_key)
        if target_key:
            mapped_row[target_key] = value
    return mapped_row


# gpt-5
def _parse_csv_import(content: bytes, mapping: dict[str, str]) -> list[dict[str, Any]]:
    stream = io.StringIO(content.decode("utf-8-sig"))
    reader = csv.DictReader(stream)
    return [_apply_import_mapping(_clean_import_row(row), mapping) for row in reader]


# gpt-5
def _load_openpyxl_workbook(content: bytes):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ValidationException("XLSX import requires openpyxl to be installed") from exc
    return load_workbook(io.BytesIO(content), read_only=True, data_only=True)


# gpt-5
def _parse_xlsx_import(content: bytes, mapping: dict[str, str]) -> list[dict[str, Any]]:
    workbook = _load_openpyxl_workbook(content)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers_row = next(rows, None)
        if headers_row is None:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in headers_row]
        parsed_rows: list[dict[str, Any]] = []
        for values in rows:
            raw_row = {
                headers[index]: values[index]
                for index in range(len(headers))
                if headers[index]
            }
            parsed_rows.append(_apply_import_mapping(_clean_import_row(raw_row), mapping))
        return parsed_rows
    finally:
        workbook.close()


# gpt-5
def _parse_import_file(file: UploadFile, mapping: Optional[dict]) -> list[dict[str, Any]]:
    extension = Path(file.filename or "").suffix.lower()
    if extension not in ALLOWED_IMPORT_EXTENSIONS:
        raise ValidationException("Only CSV and XLSX files are supported")
    if file.content_type and file.content_type not in ALLOWED_IMPORT_TYPES:
        raise ValidationException("Only CSV and XLSX files are supported")

    parsed_mapping = _normalize_import_mapping(mapping)
    content = file.file.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise ValidationException("Import file is too large")

    if extension == ".csv":
        return _parse_csv_import(content, parsed_mapping)
    if extension == ".xlsx":
        return _parse_xlsx_import(content, parsed_mapping)
    raise ValidationException("Unsupported import file format")


# gpt-5
def _parse_import_mapping_param(mapping: Optional[str]) -> Optional[dict]:
    if not mapping:
        return None
    try:
        parsed_mapping = json.loads(mapping)
    except (json.JSONDecodeError, TypeError):
        raise ValidationException("Invalid mapping format")
    return _normalize_import_mapping(parsed_mapping)


# gpt-5
def _extract_import_errors(exc: Exception) -> list[str]:
    if isinstance(exc, ValidationError):
        return [
            f"{'.'.join(str(part) for part in err['loc'])}: {err['msg']}"
            for err in exc.errors()
        ]
    detail = getattr(exc, "detail", None)
    if isinstance(detail, list):
        return [str(item) for item in detail]
    return [str(detail or exc)]

def register_crud_routes(router: APIRouter, model_class: Type[Any], Schema: Type[Any], PatchSchema: Type[Any], allow_public: bool, api_tag: str, write_saas_module: str = ""):
    # When the router is not blanket plan-gated (because the model is public),
    # writes still need plan enforcement. Build a per-write dependency list here.
    write_deps: list = []
    if write_saas_module:
        from ...auth.module_guard import require_module
        write_deps.append(Depends(require_module(write_saas_module)))
    
    @router.get("/metadata", response_model=dict)
    def get_metadata(
        request: Request,
        lang: Optional[str] = Query(None),
        db: Session = Depends(get_db),
        _: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=allow_public))
    ):
        """Returns metadata for dynamic GUI generation with translation support."""
        scope = _get_scope(request)
        org_id = scope.get("org_id")
        from ...aras import Aras
        view = Aras.View.get_for_model(model_class)
        if view:
            data = view.render_metadata(db=db, lang=lang)
            return ok(data, "Metadata retrieved successfully.")
        from ...logic.ui_generator import UIGenerator
        data = UIGenerator.generate_metadata(model_class, db=db, lang=lang, org_id=org_id)
        return ok(data, "Metadata retrieved successfully.")

    # claude-sonnet-4-6
    @router.get("/insights")
    def get_insights(
        request: Request,
        record_id: Optional[int] = Query(None),
        db: Session = Depends(get_db),
        _: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=allow_public))
    ):
        """Returns summary metrics (insights) for this resource or a specific record."""
        # record_id -> form mode: use __form_insights__ scoped to that row
        # no record_id -> list mode: use __insights__ scoped to org
        if record_id is not None:
            raw = getattr(model_class, "__form_insights__", []) or []
        else:
            raw = getattr(model_class, "__insights__", []) or []

        org_id = _get_scope(request).get("org_id")
        table = model_class.__table__
        results = []
        for ins in raw:
            try:
                # gemini-3-flash-preview: Refactored to avoid raw SQL string interpolation.
                if "sql" in ins:
                    # Deprecated: Log a warning and use text() safely with bound params
                    logging.warning(f"Insight 'sql' key is deprecated in {model_class.__tablename__}. Use ORM expressions.")
                    params = {}
                    if record_id is not None: params["record_id"] = record_id
                    if org_id: params["org_id"] = org_id
                    val = db.execute(text(ins["sql"]), params).scalar()
                else:
                    agg_key = ins.get("agg", "count")
                    col_name = ins.get("field", "id")
                    
                    # Map common aggregates to SQLAlchemy functions
                    agg_map = {"sum": func.sum, "avg": func.avg, "max": func.max, "min": func.min, "count": func.count}
                    agg_fn = agg_map.get(agg_key, func.count)
                    
                    target_col = getattr(table.c, col_name) if hasattr(table.c, col_name) else table.c.id
                    stmt = select(agg_fn(target_col)).select_from(table)
                    
                    if record_id is not None:
                        stmt = stmt.where(table.c.id == record_id)
                    elif org_id and hasattr(model_class, "org_id"):
                        stmt = stmt.where(table.c.org_id == org_id)
                    val = db.execute(stmt).scalar()

                results.append({
                    "key": ins.get("key", ins.get("label", "").lower().replace(" ", "_")),
                    "label": ins.get("label", ""),
                    "value": float(val) if val is not None else 0,
                    "format": ins.get("format", "number"),
                    "icon": ins.get("icon", None),
                    "prefix": ins.get("prefix", None),
                    "suffix": ins.get("suffix", None),
                })
            except Exception as e:
                logging.warning(f"Insight eval failed for {model_class.__tablename__}: {e}")
        return ok(results, "Insights retrieved.")

    @router.get("/")
    def list_items_slashed(
        request: Request,
        page: int = Query(1, ge=1),
        per_page: int = Query(20, ge=1, le=MAX_PER_PAGE),
        search: Optional[str] = None,
        filters: Optional[str] = None,
        order_by: Optional[str] = None,
        desc: bool = True,
        db: Session = Depends(get_db),
        user: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=allow_public))
    ):
        return list_items(request, page, per_page, search, filters, order_by, desc, db, user)

    @router.get("")
    def list_items(
        request: Request,
        page: int = Query(1, ge=1),
        per_page: int = Query(20, ge=1, le=MAX_PER_PAGE),
        search: Optional[str] = None,
        filters: Optional[str] = None,
        order_by: Optional[str] = None,
        desc: bool = True,
        db: Session = Depends(get_db),
        user: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=allow_public))
    ):
        """Lists records with pagination, filtering, and search."""
        parsed_filters = None
        if filters:
            try: parsed_filters = json.loads(filters)
            except (json.JSONDecodeError, TypeError): raise ValidationException("Invalid filters format. Must be JSON.")
        parsed_filters = _apply_scope_filters(model_class, request, list(parsed_filters or []))
        paginated_data = model_class.paginate(db, page=page, per_page=per_page, search=search, filters=parsed_filters, order_by=order_by, desc=desc)
        return ok(paginated_data, "Items listed successfully.")

    @router.get("/export")
    def export_items(
        request: Request,
        search: Optional[str] = None,
        filters: Optional[str] = None,
        order_by: Optional[str] = None,
        desc: bool = True,
        db: Session = Depends(get_db),
        _: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=allow_public))
    ):
        """Exports records to a CSV file based on current filters and search."""
        parsed_filters = None
        if filters:
            try: parsed_filters = json.loads(filters)
            except (json.JSONDecodeError, TypeError): raise ValidationException("Invalid filters format.")
        parsed_filters = _apply_scope_filters(model_class, request, list(parsed_filters or []))
        stmt = model_class._q()
        stmt = model_class.apply_filters(stmt, parsed_filters)
        stmt = model_class.apply_search(stmt, search)
        sort_col = getattr(model_class, order_by) if order_by and hasattr(model_class, order_by) else model_class.id
        stmt = stmt.order_by(sort_col.desc() if desc else sort_col.asc())

        def generate():
            output = io.StringIO()
            fieldnames = [c.name for c in model_class.get_ui_fields()]
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            yield output.getvalue()
            output.seek(0)
            for item in db.scalars(stmt).yield_per(100):
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                # Only include fields present in fieldnames to avoid DictWriter error
                row_data = {k: v for k, v in item.to_dict().items() if k in fieldnames}
                writer.writerow(row_data)
                yield output.getvalue()

        return StreamingResponse(generate(), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={model_class.__tablename__}_export.csv"})

    @router.post("/import/preview", dependencies=write_deps)
    # gpt-5
    def preview_import_items(
        request: Request,
        file: UploadFile = File(...),
        mapping: Optional[str] = Query(None),
        db: Session = Depends(get_db),
        user: Any = Depends(check_permissions(model_class.__tablename__, "CREATE"))
    ):
        """Validates import rows without committing them."""
        parsed_mapping = _parse_import_mapping_param(mapping)
        data_to_preview = _parse_import_file(file, parsed_mapping)

        rows = []
        valid_count = 0
        invalid_count = 0
        for index, row in enumerate(data_to_preview, start=1):
            payload = _inject_scope_payload(model_class, request, dict(row))
            row_result = {"row": index, "ok": True, "errors": [], "data": payload}
            savepoint = db.begin_nested()
            try:
                validated_payload = Schema(**payload).model_dump()
                model_class.create(db, validated_payload, user_id=user.id)
            except Exception as exc:
                row_result["ok"] = False
                row_result["errors"] = _extract_import_errors(exc)
                invalid_count += 1
            else:
                valid_count += 1
            finally:
                savepoint.rollback()

            if len(rows) < MAX_IMPORT_PREVIEW_ROWS:
                rows.append(row_result)

        data = {
            "total": len(data_to_preview),
            "valid": valid_count,
            "invalid": invalid_count,
            "rows": rows,
            "sample": rows[:20],
        }
        return ok(data, "Import preview generated successfully")

    @router.post("/import", dependencies=write_deps)
    # gpt-5
    def import_items(
        file: UploadFile = File(...),
        mapping: Optional[str] = Query(None),
        user: Any = Depends(check_permissions(model_class.__tablename__, "CREATE"))
    ):
        """Imports records from a CSV or XLSX file via background task."""
        parsed_mapping = _parse_import_mapping_param(mapping)
        data_to_import = _parse_import_file(file, parsed_mapping)
        from ...manager.task_manager import TaskManager
        task_id = TaskManager.enqueue_task('task_manager.import_csv_task', model_class_name=model_class.__tablename__, data=data_to_import, user_id=user.id)
        data = {"message": "Import initiated in background", "task_id": task_id}
        return ok(data, data["message"])

    @router.post("/", status_code=status.HTTP_201_CREATED, dependencies=write_deps)
    def create_item_slashed(request: Request, data: Schema, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "CREATE"))):
        return create_item(request, data, db, user)

    @router.post("", status_code=status.HTTP_201_CREATED, dependencies=write_deps)
    def create_item(request: Request, data: Schema, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "CREATE"))):
        """Creates a new record with hooks support."""
        payload = _inject_scope_payload(model_class, request, data.model_dump())
        new_item = model_class.create(db, payload, user_id=user.id)
        _save_children(db, new_item, payload, user_id=user.id)
        db.commit()
        _trigger_invalidation(db, model_class, new_item)
        _broadcast_update(model_class, "record_created", new_item.id)
        return ok(new_item.to_dict(), "Item created successfully.")

    if getattr(model_class, "__soft_delete__", False):
        @router.get("/deleted", tags=[api_tag])
        def list_deleted(request: Request, page: int = Query(1, ge=1), per_page: int = Query(20, ge=1, le=MAX_PER_PAGE), db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=False))):
            from sqlalchemy import select as sa_select
            from sqlalchemy import func
            stmt = sa_select(model_class).where(model_class.deleted_at.isnot(None))
            stmt = model_class.apply_filters(stmt, _apply_scope_filters(model_class, request, []))
            total = db.scalar(sa_select(func.count()).select_from(stmt.subquery()))
            offset = (page - 1) * per_page
            items = db.scalars(stmt.offset(offset).limit(per_page)).all()
            data = {"items": [i.to_dict() for i in items], "total": total, "page": page, "pages": (total + per_page - 1) // per_page}
            return ok(data, "Deleted items listed successfully.")

        @router.post("/{item_id}/restore")
        def restore_item(request: Request, item_id: int, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "UPDATE"))):
            from sqlalchemy import select as sa_select
            item = db.scalar(sa_select(model_class).where(model_class.id == item_id))
            if not item: raise ResourceNotFoundException("Item not found")
            _check_scope_ownership(model_class, request, item)
            if item.deleted_at is None: raise ValidationException("Record is not deleted")
            item.deleted_at = None
            item.updated_by = user.id
            db.commit()
            db.refresh(item)
            return ok(item.to_dict(), "Item restored successfully.")

    @router.get("/{item_id}/linked-documents")
    def get_linked_documents(request: Request, item_id: int, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=False))):
        item = model_class.get(db, item_id)
        if not item: raise ResourceNotFoundException("Item not found")
        _check_scope_ownership(model_class, request, item)
        docs = item.get_linked_documents(db) if hasattr(item, "get_linked_documents") else []
        return ok(docs, "Linked documents fetched.")

    @router.get("/{item_id}")
    def get_item(request: Request, item_id: int, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "READ", allow_public=allow_public))):
        item = model_class.get(db, item_id)
        if not item: raise ResourceNotFoundException("Item not found")
        _check_scope_ownership(model_class, request, item)
        res = item.to_dict()
        model_class.resolve_labels(db, [res])
        from ...base.model import Model as ArasModel
        child_defs = ArasModel._child_map.get(model_class.__tablename__, [])
        for child_def in child_defs:
            child_resource_name = child_def["resource"]
            fk_column = child_def["fk_column"]
            if fk_column and child_resource_name:
                try:
                    child_model_class = ArasModel.get_model(child_resource_name)
                    child_records = db.scalars(select(child_model_class).where(getattr(child_model_class, fk_column) == item_id)).all()
                    res[child_resource_name] = [rec.to_dict() for rec in child_records]
                    child_model_class.resolve_labels(db, res[child_resource_name])
                except KeyError: logging.warning(f"Child model {child_resource_name} not found in registry.")
                except Exception as e: logging.error(f"Error fetching child records for {child_resource_name}: {e}", exc_info=True)
        return ok(res, "Item retrieved successfully.")

    @router.put("/{item_id}", dependencies=write_deps)
    def update_item(request: Request, item_id: int, data: Schema, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "UPDATE"))):
        item = model_class.get(db, item_id)
        if not item: raise ResourceNotFoundException("Item not found")
        _check_scope_ownership(model_class, request, item)
        payload = data.model_dump(exclude_unset=True)
        item.update_self(db, payload, user_id=user.id)
        _save_children(db, item, payload, user_id=user.id)
        db.commit()
        _trigger_invalidation(db, model_class, item)
        _broadcast_update(model_class, "record_updated", item_id)
        res = item.to_dict()
        model_class.resolve_labels(db, [res])
        return ok(res, "Item updated successfully.")

    @router.patch("/{item_id}", dependencies=write_deps)
    def patch_item(request: Request, item_id: int, data: PatchSchema, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "UPDATE"))):
        item = model_class.get(db, item_id)
        if not item: raise ResourceNotFoundException("Item not found")
        _check_scope_ownership(model_class, request, item)
        payload = data.model_dump(exclude_unset=True)
        item.update_self(db, payload, user_id=user.id)
        _save_children(db, item, payload, user_id=user.id)
        db.commit()
        _trigger_invalidation(db, model_class, item)
        _broadcast_update(model_class, "record_updated", item_id)
        res = item.to_dict()
        model_class.resolve_labels(db, [res])
        return ok(res, "Item patched successfully.")

    @router.delete("/{item_id}", dependencies=write_deps)
    def delete_item(request: Request, item_id: int, db: Session = Depends(get_db), user: Any = Depends(check_permissions(model_class.__tablename__, "DELETE"))):
        item = model_class.get(db, item_id)
        if not item: raise ResourceNotFoundException("Item not found")
        _check_scope_ownership(model_class, request, item)
        item.delete_self(db, user_id=user.id)
        db.commit()
        _trigger_invalidation(db, model_class, item)
        _broadcast_update(model_class, "record_deleted", item_id)
        return ok({"id": item_id}, "Item deleted successfully.")
